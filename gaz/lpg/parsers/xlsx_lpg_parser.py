import openpyxl

def XLSXLpgParser(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    rows = 99
    parsed = {'date': [],
    'mileage_total': [],
    'price': [],
    'volume': [],
    'benz_price': [],
    'cost': [],
    'mileage': [],
    'consump': [],
    'saving': [],}
    
    for i in range(2, rows):
        date = sheet.cell(row = i, column = 1)
        mileage_total = sheet.cell(row = i, column = 2)
        price = sheet.cell(row = i, column = 3)
        volume = sheet.cell(row = i, column = 4)
        benz_price = sheet.cell(row = i, column = 5)
        cost = sheet.cell(row = i, column = 6)
        mileage = sheet.cell(row = i, column = 7)
        consump = sheet.cell(row = i, column = 8)
        saving = sheet.cell(row = i, column = 9)
        parsed['date'].append(date.value)
        parsed['mileage_total'].append(mileage_total.value)
        parsed['price'].append(price.value)
        parsed['volume'].append(volume.value)
        parsed['benz_price'].append(benz_price.value)
        parsed['cost'].append(round(cost.value, 2))
        parsed['mileage'].append(mileage.value)
        parsed['consump'].append(round(consump.value, 2))
        parsed['saving'].append(round(saving.value, 2))

    return parsed